from openerp import models, fields, api
from datetime import datetime
from datetime import timedelta
from openerp.report import report_sxw
from openerp.osv import osv
from openpyxl import Workbook
import csv
import sys
import os


class res_partner(models.Model):
    _inherit = 'res.partner'

    doctor = fields.Boolean('Doctor')
    patient = fields.Boolean('Patient')


class lab_test(models.Model):
    _name = 'lab.test'

    name = fields.Char(string='Name of Test')
    price = fields.Float('Price of the Test')
    hours = fields.Integer(string='Hours to preprate this Test')

    _sql_constraints = [
        ('name_unique',
         'UNIQUE(name)',
         "The Name must be Unique"),
    ]


class lab_order(models.Model):
    _name = 'lab.order'


    partner_id = fields.Many2one('res.partner', string='Patient')
    doctor_id = fields.Many2one('res.partner', string='Ref. Doctor')
    date = fields.Datetime(default=fields.Datetime.now)
    dob = fields.Date('Date of Birth', required=True)
    gender = fields.Selection([('male', 'Male'), ('female', 'Female')], required=True)
    line_ids = fields.One2many('lab.order.line', 'order_id', 'Lines')
    color = fields.Integer()
    state = fields.Selection([('new', "New"), ('sample_received', "Sample Received"), ('complete', "Complete"), ('collected', "Collected"), ], default='new')

    age = fields.Integer(compute='_age', string='Age')

    def _age(self):
        for cal in self:
            cal.age = (datetime.now() - datetime.strptime(cal.dob, '%Y-%m-%d')).days / 365

    amount = fields.Float(compute='_total_amount', string='Total Price')

    def _total_amount(self):
        for lab_price in self:
            lab_price.amount = sum([line.price for line in lab_price.line_ids])

    total_hours = fields.Float(compute='_total_hours', string='Total Hours')

    def _total_hours(self):
        for lab_hours in self:
            lab_hours.total_hours = sum([line.hours for line in lab_hours.line_ids])

    @api.multi
    def action_new(self):
        self.state = 'new'

    @api.multi
    def action_received(self):
        self.state = 'sample_received'

    @api.multi
    def action_complete(self):
        self.state = 'complete'

    @api.multi
    def action_collected(self):
        self.state = 'collected'


    @api.multi
    def record_export(self):
        wb = Workbook()
        dest_filename = (r'/home/axilt/Desktop/lab_order.xlsx')
        ws = wb.worksheets[0]
        ws.title = "Lab Records"
        records = self.search([])
        ws['A1'] = 'Patient'
        ws['B1'] = 'Ref. Doctor'
        ws['C1'] = 'Date'
        ws['D1'] = 'Gender'
        row_indx = 2
        for r in records:
            ws['A' + str(row_indx)] = str(r.partner_id.name)
            ws['B' + str(row_indx)] = str(r.doctor_id.name)
            ws['C' + str(row_indx)] = str(r.date)
            ws['D' + str(row_indx)] = str(r.gender)
            row_indx += 1
        wb.save("lab_order.xlsx")
        wb.save(filename=dest_filename)


    @api.multi
    def record_import(self):
        f = open('/home/axilt/Desktop/lab_write.csv')
        csv_import = csv.reader(f)
        for row in csv_import:
            print (row)
            patient = self.env['res.partner'].create({'name': row[1]})
            doctor = self.env['res.partner'].create({'name': row[2]})
            self.create({'date':row[0],'age':row[3],'dob':row[5],'gender':row[4],'partner_id':patient.id,'doctor_id':doctor.id})
            print patient.id


class lab_order_line(models.Model):
    _name = 'lab.order.line'

    order_id = fields.Many2one('lab.order', string='Order')
    test_id = fields.Many2one('lab.test', string='Test')
    price = fields.Float('Price')
    hours = fields.Integer('Hours')
    amount = fields.Float('Total Price')
    total_hours = fields.Float('Total Hours')

    @api.onchange('test_id')
    def onchange_price(self):
        self.price = self.test_id.price

    @api.onchange('test_id')
    def onchange_hours(self):
        self.hours = self.test_id.hours